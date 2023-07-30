from os import system, name, remove
from time import sleep
import requests
from bs4 import BeautifulSoup
from numpy import zeros, int16, max as nmax, add, ndenumerate
from random import sample
from datetime import timedelta, datetime
from itertools import product
from functools import reduce
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, ConversationHandler, MessageHandler, CallbackQueryHandler, CallbackContext, ContextTypes, filters
import logging, json
import logging.handlers

CONFIRMATION = 0
ENTERING_ASSIGNATURES = 1
SELECTING_SCHEDULES = 2
EXIT = 3

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
formatter = logging.Formatter(log_format)

file_handler = logging.handlers.RotatingFileHandler('/home/sr-gus/telegram-bot-horarios/logs/logs.log')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        file_handler
    ]
)

global codes
global options
global is_option_selected
global subjects
global current_subject
global button_lists
global temp_msgs

def hour_to_interval(hour):
    hour, minutes = map(int, hour.split(':'))
    return hour * 4 + minutes // 15

async def load_html(codes, message):
    bar_length = 10
    step = len(codes) / bar_length
    days = {'Lun': 0, 'Mar': 1, 'Mie': 2, 'Jue': 3, 'Vie': 4, 'Sab': 5, 'Dom': 6}
    options = {}
    progress = 1
    progress_msg = await message.reply_text('▱'*bar_length + ' - 0 %')
    for code in codes:
        logger.info(f'Obteniendo datos para el código: {code}')
        try:
            html_text = requests.get(f'https://www.ssa.ingenieria.unam.mx/cj/tmp/programacion_horarios/{code}.html?_=1675362427735', timeout=10).text
        except requests.exceptions.Timeout:
            await message.reply_text(f'No se pudieron obtener los grupos para el código {code}.')
            bar = '▰' * round(progress / step) + '▱' * round((len(codes) - progress) / step) + ' - {:.2f} %'.format(100*progress/len(codes))
            await progress_msg.edit_text(text=bar)
            progress += 1
            continue
        soup = BeautifulSoup(html_text, 'html.parser')
        tracks = soup.find_all('td')
        groups = []
        group = []
        schedule = zeros((96, 7), dtype=int16)
        hours = []
        reading_schedule = False
        save = True
        places = 0
        for i in range(len(tracks)):
            text = tracks[i].get_text().strip()
            logger.info(f'Información para el código {code}: {text}')
            if text.isdigit() and len(group) >= 2 and int(text) < 100:
                places = int(text)
                group.append(places)
            if text == 'L+' and (len(code) != 4 or (code[0] != '5' and code[0] != '6')):
                save = False
            if ':' in text and len(group) >= 3:
                hours.append(text)
                reading_schedule = True
            elif reading_schedule:
                days_occupied = [days[day] for day in text.split(', ')]
                interval = hours[-1].split(' a ')
                hours[-1] = text + ' - ' + hours[-1]
                start, end = map(hour_to_interval, interval)
                for day in days_occupied:
                    schedule[start:end, day] = 1
                reading_schedule = False
            if (text == code and i > 0) or (i == (len(tracks)-1)):
                group.append(' / '.join(hours))
                hours = []
                group.append(schedule)
                schedule = zeros((96, 7), dtype=int16)
                if places > 0 and save:
                    groups.append(group.copy())
                    group = [text]
                save = True
            else:
                if text != 'L' and text != 'T' and ',' not in text and ':' not in text and text not in days.keys():
                    group.append(text)
        if groups:
            options[code] = groups
            await message.reply_text(f'Se obtuvieron correctamente los grupos para {code}.')
        else:
            await message.reply_text(f'No hay grupos disponibles para {code}.')
        bar = '▰' * round(progress / step) + '▱' * round((len(codes) - progress) / step) + ' - {:.2f} %'.format(100*progress/len(codes))
        await progress_msg.edit_text(text=bar)
        progress += 1
    return options

async def create_schedules(options, message):
    all_permutations = list(product(*options.values()))
    if not all_permutations:
        return []
    bar_length = 10
    step = len(all_permutations) / bar_length
    valid_permutations = []
    schedules_count = 1
    progress_msg = await message.reply_text('▱'*bar_length + ' - 0 %')
    for permutation in all_permutations:
        schedules = list((group[-1] for group in permutation))
        if nmax(reduce(add, schedules)) <= 1:
            valid_permutations.append(permutation) 
        percentage = 100*schedules_count/len(all_permutations)
        if schedules_count and (len(all_permutations) < 10 or schedules_count % (len(all_permutations)//10) == 0):
            bar = '▰' * round(schedules_count / step) + '▱' * round((len(all_permutations) - schedules_count) / step) + ' - {:.2f} %'.format(percentage)
            await progress_msg.edit_text(text=bar)
        schedules_count += 1
    await progress_msg.edit_text('▰'*bar_length + ' - 100 %')
    len_permutations = len(valid_permutations)
    if len_permutations > 0:
        await message.reply_text(f'{len_permutations} horarios creados.')
    else:
        await message.reply_text('No se pudieron formar horarios válidos.')
    return valid_permutations

async def to_xlsx(schedules, message):
    book = Workbook()
    palette = 'afeeee-96c5b0-f2f3ae-90ee90-f6828c-ffdab9-bbbac6-ffb6c1-f7a9a8-edb6a3'.split('-')
    schedule_count = 1
    weekday_list = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'DOMINGO']
    bar_length = 10
    step = len(schedules) / bar_length
    progress_msg = await message.reply_text('▱'*bar_length + ' - 0 %')
    for schedule in schedules:
        if schedule_count == 1:
            sheet = book.active
            sheet.title = 'Horario {}'.format(schedule_count)
        else:
            sheet = book.create_sheet('Horario {}'.format(schedule_count))
            sheet = book['Horario {}'.format(schedule_count)]
        for i in range(len(weekday_list)):
            sheet.cell(row=2, column=i+3, value=weekday_list[i])
            sheet.cell(row=2, column=i+3).alignment = Alignment(horizontal='center', vertical='center')
        for i in range(17*4):
            start_time = datetime.strptime('07:00', '%H:%M')
            time = start_time + timedelta(minutes=15*i)
            sheet.cell(row=i+3, column=2, value=time.strftime('%H:%M'))
            sheet.cell(row=i+3, column=2).alignment = Alignment(horizontal='center', vertical='center')
        colors = sample(palette, len(schedule))
        sheet.cell(row=3, column=12, value='Clave')
        sheet.cell(row=3, column=12).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=3, column=13, value='Grupo')
        sheet.cell(row=3, column=13).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=3, column=14, value='Profesor')
        sheet.cell(row=3, column=14).alignment = Alignment(horizontal='center', vertical='center')
        for i in range(len(schedule)):
            sheet.cell(row=i+4, column=12, value=int(schedule[i][0]))
            fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
            sheet.cell(row=i+4, column=12).fill = fill
            sheet.cell(row=i+4, column=12).alignment = Alignment(horizontal='center', vertical='center')

            sheet.cell(row=i+4, column=13, value=int(schedule[i][1]))
            fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
            sheet.cell(row=i+4, column=13).fill = fill
            sheet.cell(row=i+4, column=13).alignment = Alignment(horizontal='center', vertical='center')

            sheet.cell(row=i+4, column=14, value=schedule[i][2])
            fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
            sheet.cell(row=i+4, column=14).fill = fill
            sheet.cell(row=i+4, column=14).alignment = Alignment(horizontal='center', vertical='center')

            for (row_index, column_index), value in ndenumerate(schedule[i][-1]):
                row_index = int(row_index)
                column_index = int(column_index)
                row = row_index + 3 - 28
                column = column_index + 3
                if value == 1:
                    fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
                    sheet.cell(row=row, column=column).fill = fill
        column_widths = []
        for column in sheet.columns:
            max_width = max(len(str(cell.value)) for cell in column)
            if max_width > 8:
                sheet.column_dimensions[column[0].column_letter].width = max_width*1.2
        percentage = 100*schedule_count/len(schedules)
        if len(schedules) < 10  or schedule_count % (len(schedules)//10) == 0:
            bar = '▰' * round(schedule_count / step) + '▱' * round((len(schedules) - schedule_count) / step) + ' - {:.2f} %'.format(percentage)
            await progress_msg.edit_text(text=bar)
        schedule_count += 1
    await progress_msg.edit_text('▰'*bar_length + ' - 100 %')
    book.save(filename = f'{message.chat_id}.xlsx')
    await message.reply_text('Tus horarios han sido exportados a Excel exitosamente.')
    await message.reply_document(document=open(f'{message.chat_id}.xlsx', 'rb'))
    remove(f'{message.chat_id}.xlsx')

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text('Usa el comando /iniciar para iniciar a usar el bot.')

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    username = update.message.chat.username
    if username != None:
        await update.message.reply_text(f'Hola {update.message.chat.username}!!!')
    await update.message.reply_text('Bienvenid@ al bot de horarios para la FI.')
    await update.message.reply_text('Ingresa los códigos de tus materias separados por comas.')
    return ENTERING_ASSIGNATURES

async def handle_codes(update: Update, _: CallbackContext) -> int:
    global codes
    codes = [code.strip() for code in update.message.text.split(',')]
    codes = [code for code in codes if code != '']
    if all([item.isdigit() for item in codes]):
        for i in range(len(codes)):
            if codes[i][0] == '0':
                codes[i] = codes[i][1:]
        reply_keyboard = [['Obtener grupos', 'Repetir selección']]
        await update.message.reply_text(
            'Se obtendrán los grupos de las materias con los siguientes códigos:\n- ' + '\n- '.join(codes), 
            reply_markup=ReplyKeyboardMarkup(
                reply_keyboard, 
                one_time_keyboard=True, 
                input_field_placeholder='¿Es correcto?'
            )
        )
        return CONFIRMATION
    else:
        await update.message.reply_text('Entrada inválida, ingresa solo números separados por comas.')
        return ENTERING_ASSIGNATURES

async def handle_schedules(update: Update, _: CallbackContext) -> None:
    global codes
    global options
    global is_option_selected
    global subjects
    global current_subject
    global button_lists
    global temp_msgs
   
    query = update.callback_query
    await query.answer()
    selected_option = query.data
    
    if selected_option == 'Aceptar':
        for msg in temp_msgs:
            await msg.delete()
        if subjects[current_subject] == subjects[-1]:
            for subject in subjects:
                list_of_options = options[subject]
                len_options = len(list_of_options)
                new_list_of_options = []
                for i in range(len_options):
                    if is_option_selected[subject][list_of_options[i][1]]:
                        new_list_of_options.append(list_of_options[i])
                options[subject] = new_list_of_options
            await query.message.reply_text('Generando horarios a partir de las opciones seleccionadas.', reply_markup=None)
            schedules = await create_schedules(options, query.message)
            if schedules:
                await query.message.reply_text('Exportando horarios a Excel.', reply_markup=None)
                await to_xlsx(schedules, query.message)
            return EXIT
        else:
            for msg in temp_msgs:
                try:
                    await msg.delete()
                except:
                    pass
            temp_msgs = []
            count_options = 0
            current_subject += 1
            button_list = []
            text = f'GRUPOS DISPONIBLES PARA {subjects[current_subject]}\n' + '-'*10 
            for option in options[subjects[current_subject]]:
                cupo = 0
                for element in option:
                    if isinstance(element, int) and element < 100:
                        cupo = element
                text += f'\n\nGRUPO {option[1]}\nProfesor: {option[2]}\nHorario: {option[-2]}'
                button_list.append([InlineKeyboardButton(f'GRUPO {option[1]} | {cupo} ☐', callback_data=option[1])])
                count_options += 1
                if count_options % 10 == 0 or count_options == len(options[subjects[current_subject]]):
                    temp_msgs.append(await query.message.reply_text(text))
                    text = ''
            button_list.append([InlineKeyboardButton('Aceptar', callback_data='Aceptar')])
            button_lists[subjects[current_subject]] = button_list
            reply_markup = InlineKeyboardMarkup(button_list)
            temp_msgs.append(await query.message.reply_text('SELECCIONA TUS GRUPOS: ', reply_markup=reply_markup))
    else:
        button_list = []
        for option in options[subjects[current_subject]]:
            if option[1] == selected_option:
                value = is_option_selected[subjects[current_subject]][selected_option]
                is_option_selected[subjects[current_subject]][selected_option] = not value
            for element in option:
                if isinstance(element, int) and element < 100:
                    cupo = element
            if is_option_selected[subjects[current_subject]][option[1]]:
                button_list.append([InlineKeyboardButton(f'GRUPO {option[1]} | {cupo} ☑', callback_data=option[1])])
            else:
                button_list.append([InlineKeyboardButton(f'GRUPO {option[1]} | {cupo} ☐', callback_data=option[1])])
        button_list.append([InlineKeyboardButton('Aceptar', callback_data='Aceptar')])
        button_lists[subjects[current_subject]] = button_list
        reply_markup = InlineKeyboardMarkup(button_list)
        await query.message.edit_text('SELECCIONA TUS GRUPOS: ', reply_markup=reply_markup)
    return SELECTING_SCHEDULES

async def handle_confirmation(update: Update, _: CallbackContext) -> int:
    global codes
    global options
    global is_option_selected
    global subjects
    global current_subject
    global button_lists
    global temp_msgs
 
    if update.message.text == 'Obtener grupos':
        global codes
        await update.message.reply_text('Obteniendo grupos para las materias seleccionadas.', reply_markup=ReplyKeyboardRemove())
        options = await load_html(codes, update.message)
        if options:
            subjects = []
            button_lists = {}
            is_option_selected = {}
            current_subject = 0
            temp_msgs = []
            for subject, list_of_options in options.items():
                subjects.append(subject)
                text = f'GRUPOS DISPONIBLES PARA {subjects[current_subject]}\n' + '-'*10
                button_list = []
                is_option_selected[subject] = {}
                count_options = 0
                for option in list_of_options:
                    cupo = 0
                    for element in option:
                        if isinstance(element, int) and element < 100:
                            cupo = element
                    if subject == subjects[current_subject]:
                        text += f'\n\nGRUPO {option[1]}\nProfesor: {option[2]}\nHorario: {option[-2]}'
                        count_options += 1
                        if count_options % 10 == 0 or count_options == len(list_of_options):
                            temp_msgs.append(await update.message.reply_text(text))
                            text = ''
                    button_list.append([InlineKeyboardButton(f'GRUPO {option[1]} | {cupo} ☐', callback_data=option[1])])
                    is_option_selected[subject][option[1]] = False
                button_list.append([InlineKeyboardButton('Aceptar', callback_data='Aceptar')])
                button_lists[subject] = button_list
            reply_markup = InlineKeyboardMarkup(button_lists[subjects[current_subject]])
            temp_msgs.append(await update.message.reply_text('SELECCIONA TUS GRUPOS: ', reply_markup=reply_markup))
            return SELECTING_SCHEDULES
        else:
            await update.message.reply_text('No hay grupos disponibles para ninguna materia.')
            await update.message.reply_text('Gracias por usar el bot.')
            return EXIT
    elif update.message.text == 'Repetir selección':
        await update.message.reply_text('Ingresa los códigos de tus materias separados por comas.', reply_markup=ReplyKeyboardRemove())
        return ENTERING_ASSIGNATURES
    else:
        pass

def main() -> None:
    global application
    global conv_handler
    application = Application.builder().token('6648405836:AAG0-vh6zU9yKdx3_K-PoYMyrKEvXYnI7yQ').build()
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ENTERING_ASSIGNATURES: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_codes)],
            SELECTING_SCHEDULES: [CallbackQueryHandler(handle_schedules)],
            CONFIRMATION: [MessageHandler(filters.Regex("^(Obtener grupos|Repetir selección)$"), handle_confirmation)],
            EXIT: [CommandHandler('start', start)],
        },
        fallbacks=[]
    )
    application.add_handler(CommandHandler('help', help_command))
    application.add_handler(conv_handler)
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()