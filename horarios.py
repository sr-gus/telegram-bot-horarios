from os import system, name
from time import sleep
from requests import get
from bs4 import BeautifulSoup
from numpy import zeros, int16, max as nmax, add, ndenumerate
from random import sample
from datetime import timedelta, datetime
from itertools import product
from functools import reduce
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

MAX_GRUPOS = 5

def clear():
    if name == 'nt':
        _ = system('cls')
    else:
        _ = system('clear')

def hour_to_interval(hour):
    hour, minutes = map(int, hour.split(':'))
    return hour * 4 + minutes // 15

def load_html(codes):
    days = {'Lun': 0, 'Mar': 1, 'Mie': 2, 'Jue': 3, 'Vie': 4, 'Sab': 5, 'Dom': 6}
    options = {}
    progress = 1
    print('  0 %')
    for code in codes:
        html_text = get('https://www.ssa.ingenieria.unam.mx/cj/tmp/programacion_horarios/{}.html?_=1675362427735'.format(code)).text
        soup = BeautifulSoup(html_text, 'html.parser')
        tracks = soup.find_all('td')
        groups = []
        group = []
        schedule = zeros((96, 7), dtype=int16)
        hours = []
        reading_schedule = False
        places = 0
        save = True
        for i in range(len(tracks)):
            text = tracks[i].get_text().strip()
            if text == 'L+' and (len(code) != 4 or (code[0] != '5' and code[0] != '6')):
                save = False
            if ':' in text:
                hours.append(text)
                reading_schedule = True
            elif reading_schedule:
                days_occupied = [days[day] for day in text.split(', ')]
                interval = hours[-1].split(' a ')
                hours[-1] = text + ' - ' + hours[-1]
                start, end = map(hour_to_interval, interval)
                for day in days_occupied:
                    schedule[start:end, day] = 1
                reading_schedule = False
            if (text == code and i > 0) or (i == (len(tracks)-1)):
                group.append(' / '.join(hours))
                hours = []
                group.append(schedule)
                schedule = zeros((96, 7), dtype=int16)
                if places > 0 and save:
                    groups.append(group.copy())
                    group = [text]
                save = True
            else:
                if text != 'L' and text != 'T' and ',' not in text and ':' not in text and text not in days.keys():
                    group.append(text)
                if text.isdigit() and len(group) >= 2:
                    places = int(text)
        if groups:
            options[code] = groups
        else:
            print('No hay grupos disponibles para {}'.format(code))
        print('  {:.2f} %'.format(100*progress/len(codes)))
        progress += 1
    return options

def select_options(options):
    for subject, list_of_options in options.items():
        selected = False
        print('Lista de grupos para {}'.format(subject))
        for i in range(len(list_of_options)):
            print('{}  |  {}  |  {}'.format(list_of_options[i][1], list_of_options[i][-2], list_of_options[i][2]))
        groups = input('\n    Ingresa los grupos de interés separados por espacios (si no ingresas ninguno se seleccionarán grupos aleatoriamente): ')
        groups = list(filter(('').__ne__, groups.split(' ')))
        if groups:
            new_list_of_options = []
            for option in list_of_options:
                if option[1] in groups:
                    new_list_of_options.append(option)
            options[subject] = new_list_of_options
        else:
            if MAX_GRUPOS > len(list_of_options):
                new_list_of_options = sample(list_of_options, len(list_of_options))
            else:
                new_list_of_options = sample(list_of_options, MAX_GRUPOS)
            options[subject] = new_list_of_options
        clear()

def create_schedules(options):
    all_permutations = list(product(*options.values()))
    if not all_permutations:
        return []
    valid_permutations = []
    schedules_count = 1
    print('  0%')
    for permutation in all_permutations:
        schedules = list((group[-1] for group in permutation))
        if nmax(reduce(add, schedules)) <= 1:
            valid_permutations.append(permutation) 
        percentage = 100*schedules_count/len(all_permutations)
        if schedules_count and (len(all_permutations) < 10 or schedules_count % (len(all_permutations)//10) == 0):
            print('  {:.2f} %'.format(percentage))
        schedules_count += 1
    print(' {} horarios creados'.format(len(valid_permutations)))
    return valid_permutations

def to_xlsx(schedules):
    book = Workbook()
    palette = 'afeeee-96c5b0-f2f3ae-90ee90-f6828c-ffdab9-bbbac6-ffb6c1-f7a9a8-edb6a3'.split('-')
    schedule_count = 1
    weekday_list = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'DOMINGO']
    print('  0 %')
    files = 1
    for schedule in schedules:
        if schedule_count == 1:
            sheet = book.active
            sheet.title = 'Horario {}'.format(schedule_count)
        else:
            sheet = book.create_sheet('Horario {}'.format(schedule_count))
            sheet = book['Horario {}'.format(schedule_count)]
        for i in range(len(weekday_list)):
            sheet.cell(row=2, column=i+3, value=weekday_list[i])
            sheet.cell(row=2, column=i+3).alignment = Alignment(horizontal='center', vertical='center')
        for i in range(17*4):
            start_time = datetime.strptime('07:00', '%H:%M')
            time = start_time + timedelta(minutes=15*i)
            sheet.cell(row=i+3, column=2, value=time.strftime('%H:%M'))
            sheet.cell(row=i+3, column=2).alignment = Alignment(horizontal='center', vertical='center')
        colors = sample(palette, len(schedule))
        sheet.cell(row=3, column=12, value='Clave')
        sheet.cell(row=3, column=12).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=3, column=13, value='Grupo')
        sheet.cell(row=3, column=13).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=3, column=14, value='Profesor')
        sheet.cell(row=3, column=14).alignment = Alignment(horizontal='center', vertical='center')
        for i in range(len(schedule)):
            sheet.cell(row=i+4, column=12, value=int(schedule[i][0]))
            fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
            sheet.cell(row=i+4, column=12).fill = fill
            sheet.cell(row=i+4, column=12).alignment = Alignment(horizontal='center', vertical='center')

            sheet.cell(row=i+4, column=13, value=int(schedule[i][1]))
            fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
            sheet.cell(row=i+4, column=13).fill = fill
            sheet.cell(row=i+4, column=13).alignment = Alignment(horizontal='center', vertical='center')

            sheet.cell(row=i+4, column=14, value=schedule[i][2])
            fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
            sheet.cell(row=i+4, column=14).fill = fill
            sheet.cell(row=i+4, column=14).alignment = Alignment(horizontal='center', vertical='center')

            for (row_index, column_index), value in ndenumerate(schedule[i][-1]):
                row_index = int(row_index)
                column_index = int(column_index)
                row = row_index + 3 - 28
                column = column_index + 3
                if value == 1:
                    fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
                    sheet.cell(row=row, column=column).fill = fill
        column_widths = []
        for column in sheet.columns:
            max_width = max(len(str(cell.value)) for cell in column)
            if max_width > 8:
                sheet.column_dimensions[column[0].column_letter].width = max_width*1.2
        percentage = 100*schedule_count/len(schedules)
        if len(schedules) < 10  or schedule_count % (len(schedules)//10) == 0:
            print('  {:.2f} %'.format(percentage))
        if schedule_count % 100 == 0:
            book.save(filename = 'horarios {} - {}.xlsx'.format(100*(files-1)+1, 100*files))
            book = Workbook()
            sheet = book.active
            sheet.title = 'Horario {}'.format(schedule_count+1)
            files += 1
        schedule_count += 1 
    if schedule_count % 100 != 0:
        book.save(filename = 'horarios {} - {}.xlsx'.format(100*(files-1)+1, schedule_count-1))

if __name__ == '__main__':
    clear()
    codes = input('Ingresa las claves de tus materias separadas por espacios: ').split(' ')
    print(codes)
    sleep(3)
    print('Cargando grupos de la página de la SSA...')
    options = load_html(codes)
    clear()
    select_options(options)
    clear()
    print('Creando horarios...')
    schedules = create_schedules(options)
    clear()
    if schedules:
        print('Exportando horarios a XLSX...')
        to_xlsx(schedules)
        clear()
        print('Horarios creados y exportados')
    else:
        print('No se pudieron formar horarios válidos con las opciones deseadas')
    sleep(5)
